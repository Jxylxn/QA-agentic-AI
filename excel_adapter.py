"""
Excel & Google Sheets Dual-Mode Adapter
Handles both local Excel and Google Sheets seamlessly
"""

import openpyxl
import gspread
from pathlib import Path
from typing import Optional, Dict, List

TC_COLUMNS = {
    "feature": 0,
    "function": 1,
    "sub_function": 2,
    "tc_ref": 4,
    "title": 5,
    "platform": 6,
    "preconditions": 8,
    "test_data": 9,
    "steps": 10,
    "expected_results": 11,
    "test_path": 12,
    "priority": 13,
    "complexity": 14,
    "regression": 15,
    "to_be_automate": 16,
    "test_design_status": 17,
    "reviewer": 18,
    "remark": 23,
    "jira_description": 24,
}

# Master TC header row — must match exactly the column indices above
MASTER_TC_HEADERS = [
    "Feature",            # 0
    "Function",           # 1
    "Sub function",       # 2
    "Jira ref",           # 3
    "TC ref",             # 4
    "Title",              # 5
    "Platform",           # 6
    "App ver.",           # 7
    "Preconditions",      # 8
    "Test data",          # 9
    "Step",               # 10
    "Expected result",    # 11
    "Test path",          # 12
    "Priority",           # 13
    "Complexity",         # 14
    "Regression",         # 15
    "To be automate",     # 16
    "Test Design Status", # 17
    "Reviewer",           # 18
    "Note",               # 19
    "Test result",        # 20
    "Tester",             # 21
    "Attached file",      # 22
    "Remark",             # 23
    "JIRA Description",   # 24
]


class ExcelAdapter:
    """Unified adapter for Excel and Google Sheets"""

    def __init__(self, mode: str = "excel", excel_path: Optional[Path] = None,
                 google_sheet_id: Optional[str] = None, credentials_path: Optional[Path] = None):
        """
        mode: "excel" or "google"
        excel_path: Path to Excel file (for excel mode)
        google_sheet_id: Google Sheet ID (for google mode)
        credentials_path: Path to credentials.json (for google mode)
        """
        self.mode = mode
        self.excel_path = excel_path
        self.google_sheet_id = google_sheet_id
        self.credentials_path = credentials_path
        self.gc = None

        if self.mode == "google" and credentials_path:
            try:
                self.gc = gspread.service_account(filename=str(credentials_path))
            except Exception as e:
                raise Exception(f"Failed to authenticate Google Sheets: {e}")

    def get_last_tc_ref(self, sheet: str) -> int:
        """Get the last TC reference number in a sheet"""
        try:
            if self.mode == "excel":
                return self._excel_get_last_tc_ref(sheet)
            else:
                return self._google_get_last_tc_ref(sheet)
        except Exception as e:
            print(f"Error getting last TC ref: {e}")
            return 0

    def _excel_get_last_tc_ref(self, sheet: str) -> int:
        """Get last TC ref from Excel"""
        if not self.excel_path or not self.excel_path.exists():
            return 0

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            if sheet not in wb.sheetnames:
                return 0

            ws = wb[sheet]
            last_ref = 0
            for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
                tc_ref = row[4] if len(row) > 4 else None
                if tc_ref is not None:
                    try:
                        last_ref = max(last_ref, int(str(tc_ref).strip()))
                    except (ValueError, TypeError):
                        pass
            wb.close()
            return last_ref
        except Exception as e:
            print(f"Excel error: {e}")
            return 0

    def _google_get_last_tc_ref(self, sheet: str) -> int:
        """Get last TC ref from Google Sheets"""
        if not self.gc or not self.google_sheet_id:
            return 0

        try:
            sh = self.gc.open_by_key(self.google_sheet_id)
            ws = sh.worksheet(sheet)

            # Get all values (skip header)
            records = ws.get_all_records()
            last_ref = 0

            for record in records:
                tc_ref = record.get("tc_ref", "")
                if tc_ref:
                    try:
                        last_ref = max(last_ref, int(str(tc_ref).strip()))
                    except (ValueError, TypeError):
                        pass
            return last_ref
        except Exception as e:
            print(f"Google Sheets error: {e}")
            return 0

    def write_tcs(self, sheet: str, tcs: List[Dict], create_new: bool = False,
                  new_sheet_name: Optional[str] = None, starting_ref: int = 1) -> Dict:
        """Write TCs to Excel or Google Sheets"""
        try:
            if self.mode == "excel":
                return self._excel_write_tcs(sheet, tcs, create_new, new_sheet_name, starting_ref)
            else:
                return self._google_write_tcs(sheet, tcs, create_new, new_sheet_name, starting_ref)
        except Exception as e:
            raise Exception(f"Failed to write TCs: {e}")

    def _excel_write_tcs(self, sheet: str, tcs: List[Dict], create_new: bool = False,
                         new_sheet_name: Optional[str] = None, starting_ref: int = 1) -> Dict:
        """Write TCs to Excel"""
        if not self.excel_path or not self.excel_path.exists():
            raise Exception(f"Excel file not found at {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path)

        if create_new:
            if not new_sheet_name:
                raise Exception("Sheet name required for new sheet creation")
            if new_sheet_name in wb.sheetnames:
                raise Exception(f"Sheet '{new_sheet_name}' already exists")

            template_ws = wb.worksheets[0]
            ws = wb.copy_worksheet(template_ws)
            ws.title = new_sheet_name

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            target_sheet = new_sheet_name
            last_row = 1
        else:
            if sheet not in wb.sheetnames:
                raise Exception(f"Sheet '{sheet}' not found")
            ws = wb[sheet]
            target_sheet = sheet
            last_row = ws.max_row
            while last_row > 1:
                row_vals = [ws.cell(row=last_row, column=c + 1).value for c in range(20)]
                if any(v is not None for v in row_vals):
                    break
                last_row -= 1

        rows_added = 0
        for idx, tc in enumerate(tcs):
            if create_new and "tc_ref" in tc:
                tc["tc_ref"] = starting_ref + idx

            last_row += 1
            for field, col_idx in TC_COLUMNS.items():
                value = tc.get(field, "")
                ws.cell(row=last_row, column=col_idx + 1).value = value
            rows_added += 1

        wb.save(self.excel_path)
        wb.close()

        return {
            "success": True,
            "rows_added": rows_added,
            "sheet": target_sheet,
            "is_new_sheet": create_new,
            "source": "excel"
        }

    def _google_write_tcs(self, sheet: str, tcs: List[Dict], create_new: bool = False,
                          new_sheet_name: Optional[str] = None, starting_ref: int = 1) -> Dict:
        """Write TCs to Google Sheets"""
        if not self.gc or not self.google_sheet_id:
            raise Exception("Google Sheets not configured")

        sh = self.gc.open_by_key(self.google_sheet_id)

        if create_new:
            if not new_sheet_name:
                raise Exception("Sheet name required for new sheet creation")

            existing_sheets = [ws.title for ws in sh.worksheets()]
            if new_sheet_name in existing_sheets:
                raise Exception(f"Sheet '{new_sheet_name}' already exists")

            # Create new sheet with proper Master TC headers
            ws = sh.add_worksheet(title=new_sheet_name, rows=1000, cols=30)

            # Write proper Master TC header row
            ws.append_row(MASTER_TC_HEADERS)

            # Apply full Master TC formatting via batch update
            try:
                self._apply_google_sheet_formatting(sh, ws)
            except Exception as fmt_err:
                print(f"Warning: Could not apply sheet formatting: {fmt_err}")

            target_sheet = new_sheet_name
            # start_row is 2 (after header)
            existing_rows = 1
        else:
            try:
                ws = sh.worksheet(sheet)
            except gspread.exceptions.WorksheetNotFound:
                raise Exception(f"Sheet '{sheet}' not found")

            target_sheet = sheet
            values = ws.get_all_values()
            existing_rows = len(values)

        rows_to_append = []
        for idx, tc in enumerate(tcs):
            if create_new and "tc_ref" in tc:
                tc["tc_ref"] = starting_ref + idx

            # Row number in the sheet (1-indexed): header=1, first data=2
            row_num = existing_rows + idx + 1

            # Build row according to column mapping (25 columns)
            row = [""] * 25
            for field, col_idx in TC_COLUMNS.items():
                if field == "jira_description":
                    # Use CONCATENATE formula with blank lines between sections
                    # Each section: label + CHAR(10) + value + CHAR(10) + CHAR(10) (blank line)
                    row[col_idx] = (
                        f'=CONCATENATE('
                        f'"Pre-condition :",CHAR(10),'
                        f'I{row_num},'
                        f'CHAR(10),CHAR(10),'
                        f'"Test Data :",CHAR(10),'
                        f'J{row_num},'
                        f'CHAR(10),CHAR(10),'
                        f'"Test Step :",CHAR(10),'
                        f'K{row_num},'
                        f'CHAR(10),CHAR(10),'
                        f'"Expected Result :",CHAR(10),'
                        f'L{row_num},'
                        f'CHAR(10),CHAR(10),'
                        f'"Remark :",CHAR(10),'
                        f'X{row_num})'
                    )
                else:
                    value = tc.get(field, "")
                    row[col_idx] = str(value) if value else ""

            rows_to_append.append(row)

        # Append all rows (use USER_ENTERED so formulas are evaluated)
        if rows_to_append:
            ws.append_rows(rows_to_append, value_input_option="USER_ENTERED")

        return {
            "success": True,
            "rows_added": len(rows_to_append),
            "sheet": target_sheet,
            "is_new_sheet": create_new,
            "source": "google"
        }

    def _apply_google_sheet_formatting(self, sh, ws) -> None:
        """Apply Master TC formatting to a newly created Google Sheet worksheet.
        Includes colored chip-style dropdowns and proper row/cell formatting.
        """
        sheet_id = ws.id
        requests = []

        # ── Helper to build color dict ────────────────────────────────────────
        def rgb(r, g, b):
            return {"red": r, "green": g, "blue": b}

        # ── 1. Header row: black bg + white bold text ─────────────────────────
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0, "endRowIndex": 1,
                    "startColumnIndex": 0, "endColumnIndex": 25
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": rgb(0.0, 0.0, 0.0),
                        "textFormat": {
                            "bold": True,
                            "foregroundColor": rgb(1.0, 1.0, 1.0)
                        },
                        "wrapStrategy": "WRAP",
                        "verticalAlignment": "MIDDLE"
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,wrapStrategy,verticalAlignment)"
            }
        })

        # ── 2. Data rows: wrap text + top alignment ───────────────────────────
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1, "endRowIndex": 1000,
                    "startColumnIndex": 0, "endColumnIndex": 25
                },
                "cell": {
                    "userEnteredFormat": {
                        "wrapStrategy": "WRAP",
                        "verticalAlignment": "TOP"
                    }
                },
                "fields": "userEnteredFormat(wrapStrategy,verticalAlignment)"
            }
        })

        # ── 3. All rows: 21px height ──────────────────────────────────────────
        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": 0,
                    "endIndex": 1000
                },
                "properties": {"pixelSize": 21},
                "fields": "pixelSize"
            }
        })

        # ── 4. Dropdown validation + conditional format colors ───────────────
        CHIP_WHITE = rgb(1.0, 1.0, 1.0)
        CHIP_DARK  = rgb(0.12, 0.12, 0.12)
        CHIP_BROWN = rgb(0.28, 0.20, 0.0)

        # col_idx → {value: (bg_rgb, text_rgb)}
        dropdown_color_specs = {
            6: {   # G – Platform
                "TH":      (rgb(0.26, 0.52, 0.96), CHIP_WHITE),   # blue
                "GL":      (rgb(0.13, 0.70, 0.45), CHIP_WHITE),   # green
                "GL & TH": (rgb(0.63, 0.44, 0.94), CHIP_WHITE),   # purple
            },
            12: {  # M – Test path
                "Positive": (rgb(0.20, 0.78, 0.35), CHIP_WHITE),  # green
                "Negative": (rgb(0.90, 0.24, 0.19), CHIP_WHITE),  # red
            },
            13: {  # N – Priority
                "High":   (rgb(0.90, 0.24, 0.19), CHIP_WHITE),    # red
                "Medium": (rgb(0.98, 0.73, 0.01), CHIP_BROWN),    # yellow
                "Low":    (rgb(0.84, 0.86, 0.88), CHIP_DARK),     # gray
            },
            14: {  # O – Complexity
                "High":   (rgb(0.96, 0.48, 0.0),  CHIP_WHITE),    # orange
                "Medium": (rgb(0.98, 0.73, 0.01), CHIP_BROWN),    # yellow
                "Low":    (rgb(0.84, 0.86, 0.88), CHIP_DARK),     # gray
            },
            15: {  # P – Regression
                "Yes": (rgb(0.20, 0.78, 0.35), CHIP_WHITE),       # green
                "No":  (rgb(0.84, 0.86, 0.88), CHIP_DARK),        # gray
            },
            16: {  # Q – To be automate
                "Yes": (rgb(0.26, 0.52, 0.96), CHIP_WHITE),       # blue
                "No":  (rgb(0.84, 0.86, 0.88), CHIP_DARK),        # gray
            },
            17: {  # R – Test Design Status
                "Reviewed":        (rgb(0.11, 0.47, 0.24), CHIP_WHITE),  # dark green
                "Ready To Review": (rgb(0.26, 0.52, 0.96), CHIP_WHITE),  # blue
                "Draft":           (rgb(0.98, 0.73, 0.01), CHIP_BROWN),  # yellow
            },
            18: {  # S – Reviewer
                "Tang": (rgb(0.63, 0.44, 0.94), CHIP_WHITE),      # purple
                "TT":   (rgb(0.13, 0.70, 0.67), CHIP_WHITE),      # teal
            },
        }

        for col_idx, value_colors in dropdown_color_specs.items():
            dropdown_values = list(value_colors.keys())

            requests.append({
                "setDataValidation": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 1, "endRowIndex": 1000,
                        "startColumnIndex": col_idx, "endColumnIndex": col_idx + 1
                    },
                    "rule": {
                        "condition": {
                            "type": "ONE_OF_LIST",
                            "values": [{"userEnteredValue": v} for v in dropdown_values]
                        },
                        "showCustomUi": True,
                        "strict": False
                    }
                }
            })

            for value, (bg_color, text_color) in value_colors.items():
                requests.append({
                    "addConditionalFormatRule": {
                        "rule": {
                            "ranges": [{
                                "sheetId": sheet_id,
                                "startRowIndex": 1, "endRowIndex": 1000,
                                "startColumnIndex": col_idx,
                                "endColumnIndex": col_idx + 1
                            }],
                            "booleanRule": {
                                "condition": {
                                    "type": "TEXT_EQ",
                                    "values": [{"userEnteredValue": value}]
                                },
                                "format": {
                                    "backgroundColor": bg_color,
                                    "textFormat": {
                                        "foregroundColor": text_color,
                                        "bold": True
                                    }
                                }
                            }
                        },
                        "index": 0
                    }
                })

        sh.batch_update({"requests": requests})

    def get_sheets(self) -> List[str]:
        """Get list of available sheets"""
        try:
            if self.mode == "excel":
                if not self.excel_path or not self.excel_path.exists():
                    return []
                wb = openpyxl.load_workbook(self.excel_path)
                sheets = wb.sheetnames
                wb.close()
                return sheets
            else:
                if not self.gc or not self.google_sheet_id:
                    return []
                sh = self.gc.open_by_key(self.google_sheet_id)
                return [ws.title for ws in sh.worksheets()]
        except Exception as e:
            print(f"Error getting sheets: {e}")
            return []
