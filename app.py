"""
KYC QA Agent — FastAPI Backend (Hybrid Workflow Edition)
TC generation is done externally via Claude Code chat.
This backend handles: sessions, Jira, Excel write, Figma proxy.
"""

import os
import json
import re
import uuid
import base64
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import httpx
import openpyxl
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from dotenv import load_dotenv
from excel_adapter import ExcelAdapter

load_dotenv(Path(__file__).parent / ".env", override=True)

# ─── Constants ────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
CLAUDE_MD_PATH = BASE_DIR / "CLAUDE.md"
EXCEL_PATH = BASE_DIR / "Master TC" / "Master TC KYC.xlsx"
CREDENTIALS_PATH = BASE_DIR / "credentials.json"
GOOGLE_SHEET_ID = "1C2yiPArIfee7fq92K1WJo0pMfM1D9YOYcixXF1cuolQ"
SESSIONS_DIR = BASE_DIR / "sessions"
FIGMA_FILE_KEY = "JLxd5SCV2LoPVY0DMX4D9f"

JIRA_BASE_URL = os.environ.get("JIRA_BASE_URL", "https://bitazza.atlassian.net")
JIRA_EMAIL = os.environ.get("JIRA_EMAIL", "")
JIRA_API_TOKEN = os.environ.get("JIRA_API_TOKEN", "")
JIRA_PROJECT_KEY = os.environ.get("JIRA_PROJECT_KEY", "KC")

SESSIONS_DIR.mkdir(exist_ok=True)

# Initialize adapters (will use config to switch)
excel_adapter = ExcelAdapter(mode="excel", excel_path=EXCEL_PATH)
google_adapter = ExcelAdapter(mode="google", google_sheet_id=GOOGLE_SHEET_ID,
                               credentials_path=CREDENTIALS_PATH if CREDENTIALS_PATH.exists() else None)

ENV_URLS = {
    "staging": {
        "TH Onboarding": "https://kyc.bitazzax-staging.com/th/v/th",
        "GL Onboarding": "https://kyc.bitazza-staging.com/en/v/gl",
        "EDD": "https://kyc.bitazzax-staging.com/th/v/th",
    },
    "preprod": {
        "TH Onboarding": "https://kyc.preprod-kyc.bitazza.co.th/th/v/th",
        "GL Onboarding": "https://kyc.preprod-kyc.bitazza.com/en/v/gl",
        "EDD": "https://kyc.preprod-kyc.bitazza.co.th/th/v/th",
    },
    "prod": {
        "TH Onboarding": "https://kyc.bitazza.co.th/th/v/th",
        "GL Onboarding": "https://kyc.bitazza.com/en/v/gl",
        "EDD": "https://kyc.bitazza.co.th/th/v/th",
    },
}

SHEETS = [
    "Regression_KYC", "Smoke_KYC", "E2E_KYC", "E2E_Flow", "TCM_ALL_E2E",
    "Login_Onboard", "Select_Nationality", "Term_condition", "sumsub",
    "General_info 18", "Identity_Info 28", "Address_Info 38",
    "Background_Info 48", "Disclosure_Info 58", "Financial_Info 68",
    "Fatca 78", "Additional_Docs 88", "Info_Summary", "Test_Instruction",
    "Submit_KYC", "Page_Not_Found", "KYT", "OB_Soft_Approval",
    "Multiple_Additional", "KYCmaintenance", "KYC-event Notifications",
    "EDD_Soft_Approval", "EDD_Soft_Approval_PH.2", "Search_Console",
    "Mule_Account_PH.4", "Auto Soft freeze", "API_remark_check", "Login_EDD",
]

FLOWS = ["TH Onboarding", "GL Onboarding", "EDD"]

TC_COLUMNS = {
    "feature": 0,
    "function": 1,
    "sub_function": 2,
    "tc_ref": 4,
    "title": 5,
    "platform": 6,
    "preconditions": 8,
    "test_data": 9,
    "steps": 10,
    "expected_results": 11,
    "test_path": 12,
    "priority": 13,
    "complexity": 14,
    "regression": 15,
    "to_be_automate": 16,
    "test_design_status": 17,
    "reviewer": 18,
    "remark": 23,
    "jira_description": 24,
}

# Load CLAUDE.md for the prompt template
with open(CLAUDE_MD_PATH, "r", encoding="utf-8") as _f:
    CLAUDE_MD_CONTENT = _f.read()


def _jira_auth_header() -> dict:
    token = base64.b64encode(f"{JIRA_EMAIL}:{JIRA_API_TOKEN}".encode()).decode()
    return {
        "Authorization": f"Basic {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


# ─── App ──────────────────────────────────────────────────────────────────────

app = FastAPI(title="KYC QA Agent")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


# ─── Models ───────────────────────────────────────────────────────────────────

class CreateSessionRequest(BaseModel):
    ticket_id: str
    flow: str = "TH Onboarding"
    sheet: str = ""
    environment: str = "staging"


class UpdateSessionRequest(BaseModel):
    sheet: Optional[str] = None
    flow: Optional[str] = None
    environment: Optional[str] = None
    tcs: Optional[list] = None
    tcs_written: Optional[bool] = None
    current_step: Optional[int] = None
    test_results: Optional[list] = None
    bugs_created: Optional[list] = None
    new_sheet_mode: Optional[bool] = None  # True if creating new sheet
    new_sheet_name: Optional[str] = None  # Name of new sheet
    new_sheet_starting_ref: Optional[int] = None  # Starting TC ref for new sheet
    excel_source: Optional[str] = None  # "excel" or "google"


class WriteExcelRequest(BaseModel):
    sheet: Optional[str] = None  # Existing sheet
    tcs: list[dict]
    create_new_sheet: bool = False  # Flag to create new sheet
    new_sheet_name: Optional[str] = None  # Name for new sheet
    new_sheet_starting_ref: int = 1  # Starting TC ref for new sheet


class JiraBugRequest(BaseModel):
    summary: str
    description: str
    tc_ref: Optional[str] = ""
    sheet: Optional[str] = ""
    priority: Optional[str] = "Medium"
    flow: Optional[str] = ""
    steps: Optional[str] = ""
    expected: Optional[str] = ""
    actual: Optional[str] = ""
    screenshot_url: Optional[str] = ""


class AddActivityRequest(BaseModel):
    message: str
    type: str = "info"
    link: Optional[str] = None
    link_text: Optional[str] = None


# ─── Session Helpers ──────────────────────────────────────────────────────────

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _session_path(session_id: str) -> Path:
    return SESSIONS_DIR / f"{session_id}.json"


def _load_session(session_id: str) -> dict:
    path = _session_path(session_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_session(session: dict) -> dict:
    session["updated_at"] = _now_iso()
    path = _session_path(session["id"])
    with open(path, "w", encoding="utf-8") as f:
        json.dump(session, f, ensure_ascii=False, indent=2)
    return session


def _add_activity(session: dict, message: str, type_: str = "info", link: str = "", link_text: str = "") -> None:
    entry = {
        "time": _now_iso(),
        "message": message,
        "type": type_,
    }
    if link:
        entry["link"] = link
        entry["link_text"] = link_text or link
    session.setdefault("activity", []).insert(0, entry)
    session["activity"] = session["activity"][:100]


def _adf_to_text(node) -> str:
    """Convert Atlassian Document Format to plain text."""
    if not node:
        return ""
    t = node.get("type", "")
    if t == "text":
        return node.get("text", "")
    elif t in ("paragraph", "heading", "blockquote"):
        inner = "".join(_adf_to_text(c) for c in node.get("content", []))
        return inner + "\n"
    elif t == "bulletList":
        return "".join(_adf_to_text(i) for i in node.get("content", []))
    elif t == "orderedList":
        items = node.get("content", [])
        result = ""
        for idx, item in enumerate(items, 1):
            inner = "".join(_adf_to_text(c) for c in item.get("content", []))
            result += f"{idx}. {inner.strip()}\n"
        return result
    elif t == "listItem":
        inner = "".join(_adf_to_text(c) for c in node.get("content", []))
        return f"• {inner.strip()}\n"
    elif t == "hardBreak":
        return "\n"
    elif t == "rule":
        return "\n---\n"
    elif t == "codeBlock":
        inner = "".join(_adf_to_text(c) for c in node.get("content", []))
        return f"```\n{inner}\n```\n"
    else:
        return "".join(_adf_to_text(c) for c in node.get("content", []))


def _write_tcs_to_excel_internal(sheet: str, tcs: list, create_new: bool = False, new_sheet_name: str = None, starting_ref: int = 1) -> dict:
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Master TC file not found at: {EXCEL_PATH}.")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Handle new sheet creation
    if create_new:
        if not new_sheet_name:
            raise HTTPException(status_code=400, detail="Sheet name is required for new sheet creation.")
        if new_sheet_name in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{new_sheet_name}' already exists.")

        # Copy format from first sheet (template)
        template_ws = wb.worksheets[0]
        ws = wb.copy_worksheet(template_ws)
        ws.title = new_sheet_name

        # Clear data rows, keep headers
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        target_sheet = new_sheet_name
        last_row = 1
    else:
        if sheet not in wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet}' not found.")
        ws = wb[sheet]
        target_sheet = sheet
        last_row = ws.max_row
        while last_row > 1:
            row_vals = [ws.cell(row=last_row, column=c + 1).value for c in range(20)]
            if any(v is not None for v in row_vals):
                break
            last_row -= 1

    rows_added = 0
    for idx, tc in enumerate(tcs):
        # Auto-increment tc_ref if creating new sheet
        if create_new and "tc_ref" in tc:
            tc["tc_ref"] = starting_ref + idx

        last_row += 1
        for field, col_idx in TC_COLUMNS.items():
            value = tc.get(field, "")
            ws.cell(row=last_row, column=col_idx + 1).value = value
        rows_added += 1

    wb.save(EXCEL_PATH)
    wb.close()
    return {"success": True, "rows_added": rows_added, "sheet": target_sheet, "is_new_sheet": create_new}


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = BASE_DIR / "static" / "index.html"
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@app.get("/api/config")
async def get_config():
    # Check Google Sheets availability
    google_sheets_available = False
    google_sheets_authenticated = False
    if CREDENTIALS_PATH.exists():
        google_sheets_available = True
        try:
            # Try to verify credentials work
            if google_adapter.gc is not None:
                google_sheets_authenticated = True
        except Exception:
            pass

    return {
        "sheets": SHEETS,
        "flows": FLOWS,
        "has_figma_token": bool(os.environ.get("FIGMA_ACCESS_TOKEN")),
        "excel_exists": EXCEL_PATH.exists(),
        "google_sheets_available": google_sheets_available,
        "google_sheets_authenticated": google_sheets_authenticated,
        "google_sheet_id": GOOGLE_SHEET_ID if google_sheets_available else None,
        "has_jira": bool(JIRA_EMAIL and JIRA_API_TOKEN),
        "env_urls": ENV_URLS,
        "environments": ["staging", "preprod", "prod"],
    }


# ─── Session Endpoints ────────────────────────────────────────────────────────

@app.post("/api/sessions")
async def create_session(request: CreateSessionRequest):
    session_id = str(uuid.uuid4())[:8]
    ticket_id = request.ticket_id.strip().upper()

    session = {
        "id": session_id,
        "ticket_id": ticket_id,
        "ticket_summary": "",
        "ticket_data": {},
        "sheet": request.sheet,
        "flow": request.flow,
        "environment": request.environment,
        "tcs": [],
        "tcs_written": False,
        "test_results": [],
        "bugs_created": [],
        "activity": [],
        "current_step": 1,
        "excel_source": "excel",  # Default to Excel
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }

    # Auto-fetch Jira ticket
    if ticket_id and JIRA_EMAIL and JIRA_API_TOKEN:
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    f"{JIRA_BASE_URL}/rest/api/3/issue/{ticket_id}",
                    headers=_jira_auth_header(),
                    params={"fields": "summary,description,issuetype,status,priority,assignee,labels"},
                )
            if resp.status_code == 200:
                data = resp.json()
                fields = data.get("fields", {})
                desc_raw = fields.get("description")
                description_text = ""
                if isinstance(desc_raw, dict):
                    description_text = _adf_to_text(desc_raw).strip()
                elif isinstance(desc_raw, str):
                    description_text = desc_raw.strip()

                summary = fields.get("summary", "")
                session["ticket_summary"] = summary
                session["ticket_data"] = {
                    "key": ticket_id,
                    "summary": summary,
                    "issue_type": fields.get("issuetype", {}).get("name", ""),
                    "status": fields.get("status", {}).get("name", ""),
                    "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                    "description": description_text,
                    "labels": fields.get("labels", []),
                    "url": f"{JIRA_BASE_URL}/browse/{ticket_id}",
                }
                _add_activity(session, f"Jira ticket {ticket_id} loaded: {summary[:60]}", "jira",
                              f"{JIRA_BASE_URL}/browse/{ticket_id}", ticket_id)
        except Exception as e:
            _add_activity(session, f"Failed to fetch Jira ticket: {str(e)[:80]}", "error")

    _add_activity(session, f"Session created for {ticket_id}", "info")
    _save_session(session)
    return session


@app.get("/api/sessions")
async def list_sessions():
    sessions = []
    for path in SESSIONS_DIR.glob("*.json"):
        try:
            with open(path, "r", encoding="utf-8") as f:
                s = json.load(f)
            sessions.append(s)
        except Exception:
            continue
    sessions.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    return {"sessions": sessions}


@app.get("/api/sessions/{session_id}")
async def get_session(session_id: str):
    return _load_session(session_id)


@app.patch("/api/sessions/{session_id}")
async def update_session(session_id: str, request: UpdateSessionRequest):
    session = _load_session(session_id)
    if request.sheet is not None:
        session["sheet"] = request.sheet
    if request.flow is not None:
        session["flow"] = request.flow
    if request.environment is not None:
        session["environment"] = request.environment
    if request.tcs is not None:
        session["tcs"] = request.tcs
    if request.tcs_written is not None:
        session["tcs_written"] = request.tcs_written
    if request.current_step is not None:
        session["current_step"] = request.current_step
    if request.test_results is not None:
        session["test_results"] = request.test_results
    if request.bugs_created is not None:
        session["bugs_created"] = request.bugs_created
    if request.new_sheet_mode is not None:
        session["new_sheet_mode"] = request.new_sheet_mode
    if request.new_sheet_name is not None:
        session["new_sheet_name"] = request.new_sheet_name
    if request.new_sheet_starting_ref is not None:
        session["new_sheet_starting_ref"] = request.new_sheet_starting_ref
    if request.excel_source is not None:
        session["excel_source"] = request.excel_source
        _add_activity(session, f"Excel source switched to: {request.excel_source.upper()}", "config")
    return _save_session(session)


@app.delete("/api/sessions/{session_id}")
async def delete_session(session_id: str):
    path = _session_path(session_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    path.unlink()
    return {"success": True}


@app.post("/api/sessions/{session_id}/tcs")
async def import_tcs_to_session(session_id: str, body: dict):
    session = _load_session(session_id)
    tcs = body.get("tcs", [])
    session["tcs"] = tcs
    session["tcs_written"] = False
    session["test_results"] = []
    _add_activity(session, f"Imported {len(tcs)} TC(s)", "tc")
    return _save_session(session)


@app.post("/api/sessions/{session_id}/tcs/add")
async def add_tc_to_session(session_id: str, body: dict):
    session = _load_session(session_id)
    tc = body.get("tc", {})
    session.setdefault("tcs", []).append(tc)
    session["tcs_written"] = False
    _add_activity(session, f"TC added manually: {str(tc.get('title', ''))[:50]}", "tc")
    return _save_session(session)


@app.patch("/api/sessions/{session_id}/tcs/{tc_index}")
async def update_session_tc(session_id: str, tc_index: int, body: dict):
    session = _load_session(session_id)
    tcs = session.get("tcs", [])
    if tc_index < 0 or tc_index >= len(tcs):
        raise HTTPException(status_code=404, detail=f"TC index {tc_index} out of range")
    tcs[tc_index] = body.get("tc_data", tcs[tc_index])
    session["tcs"] = tcs
    _add_activity(session, f"TC #{tc_index + 1} updated", "tc")
    return _save_session(session)


@app.delete("/api/sessions/{session_id}/tcs/{tc_index}")
async def delete_session_tc(session_id: str, tc_index: int):
    session = _load_session(session_id)
    tcs = session.get("tcs", [])
    if tc_index < 0 or tc_index >= len(tcs):
        raise HTTPException(status_code=404, detail=f"TC index {tc_index} out of range")
    removed = tcs.pop(tc_index)
    session["tcs"] = tcs
    _add_activity(session, f"TC deleted: {str(removed.get('title', ''))[:50]}", "tc")
    return _save_session(session)


@app.post("/api/sessions/{session_id}/activity")
async def add_session_activity(session_id: str, request: AddActivityRequest):
    session = _load_session(session_id)
    _add_activity(session, request.message, request.type, request.link or "", request.link_text or "")
    return _save_session(session)


@app.post("/api/sessions/{session_id}/write-excel")
async def write_session_tcs_to_excel(session_id: str):
    session = _load_session(session_id)
    tcs = session.get("tcs", [])
    if not tcs:
        raise HTTPException(status_code=400, detail="No TCs in this session")

    # Determine which source to use (default to Excel)
    excel_source = session.get("excel_source", "excel").lower()
    is_new_sheet = session.get("new_sheet_mode", False)

    try:
        if excel_source == "google":
            # Write to Google Sheets
            sheet = session.get("sheet", "")
            new_sheet_name = session.get("new_sheet_name", "")
            starting_ref = session.get("new_sheet_starting_ref", 1)

            if not google_adapter.gc:
                raise HTTPException(status_code=400, detail="Google Sheets not authenticated. Check credentials.json")

            if is_new_sheet:
                if not new_sheet_name:
                    raise HTTPException(status_code=400, detail="New sheet name is required")
                result = google_adapter.write_tcs(
                    sheet="",
                    tcs=tcs,
                    create_new=True,
                    new_sheet_name=new_sheet_name,
                    starting_ref=starting_ref
                )
                activity_msg = f"{result['rows_added']} TC(s) written to new Google Sheet '{new_sheet_name}' (starting from TC ref {starting_ref})"
            else:
                if not sheet:
                    raise HTTPException(status_code=400, detail="No sheet selected for this session")
                result = google_adapter.write_tcs(sheet=sheet, tcs=tcs, create_new=False)
                activity_msg = f"{result['rows_added']} TC(s) written to Google Sheet '{sheet}'"

        else:
            # Write to Excel (default)
            if is_new_sheet:
                new_sheet_name = session.get("new_sheet_name", "")
                starting_ref = session.get("new_sheet_starting_ref", 1)
                if not new_sheet_name:
                    raise HTTPException(status_code=400, detail="New sheet name is required")
                result = _write_tcs_to_excel_internal(
                    sheet=None,
                    tcs=tcs,
                    create_new=True,
                    new_sheet_name=new_sheet_name,
                    starting_ref=starting_ref
                )
                activity_msg = f"{result['rows_added']} TC(s) written to new Excel sheet '{new_sheet_name}' (starting from TC ref {starting_ref})"
            else:
                sheet = session.get("sheet", "")
                if not sheet:
                    raise HTTPException(status_code=400, detail="No sheet selected for this session")
                result = _write_tcs_to_excel_internal(sheet=sheet, tcs=tcs)
                activity_msg = f"{result['rows_added']} TC(s) written to Excel sheet '{sheet}'"

        session["tcs_written"] = True
        _add_activity(session, activity_msg, "excel")
        _save_session(session)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/sessions/{session_id}/test-result")
async def update_test_result(session_id: str, body: dict):
    """Mark a TC as passed or failed for manual test tracking."""
    session = _load_session(session_id)
    tc_index = body.get("tc_index")
    result = body.get("result")  # "pass" | "fail" | "skip"
    notes = body.get("notes", "")
    if tc_index is None or result not in ("pass", "fail", "skip"):
        raise HTTPException(status_code=400, detail="tc_index and valid result required")
    test_results = session.get("test_results", [])
    # Upsert
    existing = next((r for r in test_results if r.get("tc_index") == tc_index), None)
    if existing:
        existing["result"] = result
        existing["notes"] = notes
        existing["time"] = _now_iso()
    else:
        test_results.append({
            "tc_index": tc_index,
            "result": result,
            "notes": notes,
            "time": _now_iso(),
        })
    session["test_results"] = test_results
    tc_title = session.get("tcs", [{}])[tc_index].get("title", f"TC #{tc_index + 1}") if tc_index < len(session.get("tcs", [])) else f"TC #{tc_index + 1}"
    icon = {"pass": "PASS", "fail": "FAIL", "skip": "SKIP"}[result]
    _add_activity(session, f"[{icon}] {str(tc_title)[:50]}", "test")
    return _save_session(session)


@app.post("/api/sessions/{session_id}/bugs")
async def create_bug_for_session(session_id: str, body: dict):
    """Create a Jira bug linked to a session TC."""
    session = _load_session(session_id)
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        raise HTTPException(status_code=400, detail="Jira not configured")

    tc_index = body.get("tc_index")
    summary = body.get("summary", "").strip()
    priority = body.get("priority", "Medium")
    actual = body.get("actual", "")

    if not summary:
        raise HTTPException(status_code=400, detail="Summary is required")

    flow = session.get("flow", "")
    sheet = session.get("sheet", "")
    tc = session.get("tcs", [])[tc_index] if tc_index is not None and tc_index < len(session.get("tcs", [])) else {}
    steps = tc.get("steps", "")
    expected = tc.get("expected_results", "")
    tc_ref = tc.get("tc_ref", "")

    def text_node(t):
        return {"type": "text", "text": t}

    def para(*texts):
        return {"type": "paragraph", "content": [text_node(t) for t in texts]}

    def heading(t, level=3):
        return {"type": "heading", "attrs": {"level": level}, "content": [text_node(t)]}

    desc_content = []
    if flow:
        desc_content.append(para(f"Flow: {flow}"))
    if sheet:
        desc_content.append(para(f"Sheet: {sheet}"))
    if tc_ref:
        desc_content.append(para(f"TC Ref: {tc_ref}"))
    if session.get("ticket_id"):
        desc_content.append(para(f"Related Jira: {session['ticket_id']}"))

    if steps:
        desc_content.append(heading("Test Steps"))
        for line in steps.strip().split("\n"):
            if line.strip():
                desc_content.append(para(line))

    if expected:
        desc_content.append(heading("Expected Result"))
        for line in expected.strip().split("\n"):
            if line.strip():
                desc_content.append(para(line))

    if actual:
        desc_content.append(heading("Actual Result"))
        for line in actual.strip().split("\n"):
            if line.strip():
                desc_content.append(para(line))

    payload = {
        "fields": {
            "project": {"key": JIRA_PROJECT_KEY},
            "summary": summary,
            "issuetype": {"name": "Bug"},
            "priority": {"name": priority},
            "description": {"type": "doc", "version": 1, "content": desc_content},
            "labels": ["qa-agent", "auto-created"],
        }
    }

    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(
                f"{JIRA_BASE_URL}/rest/api/3/issue",
                headers=_jira_auth_header(),
                json=payload,
            )
        if resp.status_code not in (200, 201):
            raise HTTPException(status_code=resp.status_code, detail=resp.text[:500])
        data = resp.json()
        bug_key = data.get("key")
        bug_url = f"{JIRA_BASE_URL}/browse/{bug_key}"
        session.setdefault("bugs_created", []).append({
            "key": bug_key,
            "url": bug_url,
            "summary": summary,
            "tc_index": tc_index,
            "time": _now_iso(),
        })
        _add_activity(session, f"Bug {bug_key} created: {summary[:50]}", "bug", bug_url, bug_key)
        _save_session(session)
        return {"key": bug_key, "url": bug_url, "id": data.get("id")}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sessions/{session_id}/export")
async def export_session(session_id: str):
    session = _load_session(session_id)
    filename = f"session_{session_id}_{session.get('ticket_id', 'export')}.json"
    return JSONResponse(
        content=session,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/sessions/{session_id}/prompt")
async def get_session_prompt(session_id: str):
    """Generate the Claude Code prompt for TC generation based on session data."""
    session = _load_session(session_id)
    ticket_data = session.get("ticket_data", {})
    flow = session.get("flow", "TH Onboarding")
    sheet = session.get("sheet", "")

    key = ticket_data.get("key", session.get("ticket_id", ""))
    summary = ticket_data.get("summary", "")
    issue_type = ticket_data.get("issue_type", "")
    priority = ticket_data.get("priority", "")
    status = ticket_data.get("status", "")
    description = ticket_data.get("description", "")

    # Get last TC ref from Excel or Google Sheets based on session source
    last_tc_ref = 0
    excel_source = session.get("excel_source", "excel").lower()
    if sheet:
        try:
            if excel_source == "google":
                if google_adapter.gc:
                    last_tc_ref = google_adapter.get_last_tc_ref(sheet)
            else:
                # Default to Excel
                if EXCEL_PATH.exists():
                    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
                    if sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
                            tc_ref = row[4] if len(row) > 4 else None
                            if tc_ref is not None:
                                try:
                                    last_tc_ref = max(last_tc_ref, int(str(tc_ref).strip()))
                                except (ValueError, TypeError):
                                    pass
                    wb.close()
        except Exception:
            pass

    prompt = f"""จาก Jira ticket {key}:
Summary: {summary}
Type: {issue_type} | Priority: {priority} | Status: {status}

Description:
{description}

Flow: {flow}
Target Sheet: {sheet}
Last TC ref in sheet: {last_tc_ref} (new TCs start from {last_tc_ref + 1})

สร้าง TC ในรูปแบบ JSON array ตาม CLAUDE.md format:
[
  {{
    "feature": "KYC webview",
    "function": "...",
    "sub_function": "...",
    "tc_ref": {last_tc_ref + 1},
    "title": "Verify that...",
    "platform": "TH/GL/GL & TH",
    "preconditions": "...",
    "test_data": "...",
    "steps": "1. ...\\n2. ...\\n3. ...",
    "expected_results": "1. ...\\n2. ...\\n3. ...",
    "test_path": "Positive/Negative",
    "priority": "High/Medium/Low",
    "complexity": "High/Medium/Low",
    "regression": "Yes/No",
    "to_be_automate": "Yes/No",
    "test_design_status": "Ready To Review",
    "reviewer": "",
    "remark": "",
    "jira_description": "Pre-condition :\\n...\\n\\nTest Data :\\n...\\n\\nTest Step :\\n...\\n\\nExpected result :\\n..."
  }}
]

ครอบคลุม: positive paths, negative paths, validation, edge cases, platform differences (TH/GL/EDD)."""

    return {"prompt": prompt, "last_tc_ref": last_tc_ref}


# ─── Excel Endpoints ──────────────────────────────────────────────────────────

@app.get("/api/excel/last-tc-ref")
async def get_last_tc_ref(sheet: str = Query(...), source: str = Query(default="excel")):
    """Get the last TC reference from either Excel or Google Sheets.

    Parameters:
    - sheet: Sheet name
    - source: "excel" or "google" (defaults to "excel")
    """
    source = source.lower()

    if source == "google":
        try:
            if not google_adapter.gc:
                return {"last_tc_ref": 0, "error": "Google Sheets not authenticated"}
            return {"last_tc_ref": google_adapter.get_last_tc_ref(sheet), "source": "google"}
        except Exception as e:
            return {"last_tc_ref": 0, "error": str(e), "source": "google"}
    else:
        # Default to Excel
        if not EXCEL_PATH.exists():
            return {"last_tc_ref": 0, "source": "excel"}
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            if sheet not in wb.sheetnames:
                wb.close()
                return {"last_tc_ref": 0, "source": "excel"}
            ws = wb[sheet]
            last_ref = 0
            for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
                tc_ref = row[4] if len(row) > 4 else None
                if tc_ref is not None:
                    try:
                        last_ref = max(last_ref, int(str(tc_ref).strip()))
                    except (ValueError, TypeError):
                        pass
            wb.close()
            return {"last_tc_ref": last_ref, "source": "excel"}
        except Exception as e:
            return {"last_tc_ref": 0, "error": str(e), "source": "excel"}


@app.get("/api/excel/sheets")
async def get_excel_sheets(source: str = Query(default="excel")):
    """Get available sheets from either Excel or Google Sheets.

    Parameters:
    - source: "excel" or "google" (defaults to "excel")
    """
    source = source.lower()

    if source == "google":
        try:
            if not google_adapter.gc:
                return {"sheets": [], "error": "Google Sheets not authenticated", "source": "google"}
            sheets = google_adapter.get_sheets()
            return {"sheets": sheets, "source": "google"}
        except Exception as e:
            return {"sheets": [], "error": str(e), "source": "google"}
    else:
        # Default to Excel
        if not EXCEL_PATH.exists():
            return {"sheets": [], "source": "excel"}
        try:
            sheets = excel_adapter.get_sheets()
            return {"sheets": sheets, "source": "excel"}
        except Exception as e:
            return {"sheets": [], "error": str(e), "source": "excel"}


@app.post("/api/write-excel")
async def write_to_excel(request: WriteExcelRequest):
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail=f"Master TC file not found at: {EXCEL_PATH}.")
    try:
        return _write_tcs_to_excel_internal(
            sheet=request.sheet,
            tcs=request.tcs,
            create_new=request.create_new_sheet,
            new_sheet_name=request.new_sheet_name,
            starting_ref=request.new_sheet_starting_ref
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Jira Endpoints ───────────────────────────────────────────────────────────

@app.get("/api/jira/status")
async def jira_status():
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        return {"configured": False, "reason": "JIRA_EMAIL or JIRA_API_TOKEN not set"}
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                f"{JIRA_BASE_URL}/rest/api/3/myself",
                headers=_jira_auth_header(),
            )
        if resp.status_code == 200:
            data = resp.json()
            return {
                "configured": True,
                "user": data.get("displayName", ""),
                "email": data.get("emailAddress", ""),
                "project": JIRA_PROJECT_KEY,
            }
        return {"configured": False, "reason": f"HTTP {resp.status_code}"}
    except Exception as e:
        return {"configured": False, "reason": str(e)}


@app.get("/api/jira/issue/{issue_key}")
async def get_jira_issue(issue_key: str):
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        raise HTTPException(status_code=400, detail="Jira not configured")
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                f"{JIRA_BASE_URL}/rest/api/3/issue/{issue_key.upper()}",
                headers=_jira_auth_header(),
                params={"fields": "summary,description,issuetype,status,priority,assignee,labels"},
            )
        if resp.status_code == 404:
            raise HTTPException(status_code=404, detail=f"Issue {issue_key} not found")
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=resp.text[:300])
        data = resp.json()
        fields = data.get("fields", {})
        desc_raw = fields.get("description")
        description_text = ""
        if isinstance(desc_raw, dict):
            description_text = _adf_to_text(desc_raw).strip()
        elif isinstance(desc_raw, str):
            description_text = desc_raw.strip()
        return {
            "key": issue_key.upper(),
            "summary": fields.get("summary", ""),
            "issue_type": fields.get("issuetype", {}).get("name", ""),
            "status": fields.get("status", {}).get("name", ""),
            "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
            "description": description_text,
            "labels": fields.get("labels", []),
            "url": f"{JIRA_BASE_URL}/browse/{issue_key.upper()}",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/jira/recent")
async def get_recent_jira_issues(project: str = Query(default="KC"), max_results: int = Query(default=10)):
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        raise HTTPException(status_code=400, detail="Jira not configured")
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(
                f"{JIRA_BASE_URL}/rest/api/3/search/jql",
                headers=_jira_auth_header(),
                json={
                    "jql": f"project = {project} ORDER BY updated DESC",
                    "maxResults": max_results,
                    "fields": ["summary", "issuetype", "status", "priority", "updated"],
                },
            )
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=resp.text[:300])
        data = resp.json()
        issues = [
            {
                "key": i["key"],
                "summary": i["fields"].get("summary", ""),
                "type": i["fields"].get("issuetype", {}).get("name", ""),
                "status": i["fields"].get("status", {}).get("name", ""),
                "priority": i["fields"].get("priority", {}).get("name", "") if i["fields"].get("priority") else "",
                "url": f"{JIRA_BASE_URL}/browse/{i['key']}",
            }
            for i in data.get("issues", [])
        ]
        return {"issues": issues, "total": data.get("total", 0)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/jira/bug")
async def create_jira_bug(request: JiraBugRequest):
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        raise HTTPException(status_code=400, detail="Jira not configured")
    try:
        def text_node(t):
            return {"type": "text", "text": t}

        def para(*texts):
            return {"type": "paragraph", "content": [text_node(t) for t in texts]}

        def heading(t, level=3):
            return {"type": "heading", "attrs": {"level": level}, "content": [text_node(t)]}

        desc_content = []
        if request.flow:
            desc_content.append(para(f"Flow: {request.flow}"))
        if request.sheet:
            desc_content.append(para(f"Sheet: {request.sheet}"))
        if request.tc_ref:
            desc_content.append(para(f"TC Ref: {request.tc_ref}"))
        if request.steps:
            desc_content.append(heading("Test Steps"))
            for line in request.steps.strip().split("\n"):
                if line.strip():
                    desc_content.append(para(line))
        if request.expected:
            desc_content.append(heading("Expected Result"))
            for line in request.expected.strip().split("\n"):
                if line.strip():
                    desc_content.append(para(line))
        if request.actual:
            desc_content.append(heading("Actual Result"))
            for line in request.actual.strip().split("\n"):
                if line.strip():
                    desc_content.append(para(line))
        if request.description:
            desc_content.append(heading("Additional Info"))
            desc_content.append(para(request.description))

        payload = {
            "fields": {
                "project": {"key": JIRA_PROJECT_KEY},
                "summary": request.summary,
                "issuetype": {"name": "Bug"},
                "priority": {"name": request.priority or "Medium"},
                "description": {"type": "doc", "version": 1, "content": desc_content},
                "labels": ["auto-created", "qa-agent"],
            }
        }
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(
                f"{JIRA_BASE_URL}/rest/api/3/issue",
                headers=_jira_auth_header(),
                json=payload,
            )
        if resp.status_code not in (200, 201):
            raise HTTPException(status_code=resp.status_code, detail=resp.text[:500])
        data = resp.json()
        return {
            "key": data.get("key"),
            "url": f"{JIRA_BASE_URL}/browse/{data.get('key')}",
            "id": data.get("id"),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Figma Proxy ──────────────────────────────────────────────────────────────

@app.get("/api/figma/frames")
async def get_figma_frames(page_name: str = Query(...)):
    token = os.environ.get("FIGMA_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=400, detail="FIGMA_ACCESS_TOKEN not set")
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.get(
            f"https://api.figma.com/v1/files/{FIGMA_FILE_KEY}",
            headers={"X-Figma-Token": token},
        )
    data = resp.json()
    pages = data.get("document", {}).get("children", [])
    target = next((p for p in pages if p["name"].lower() == page_name.lower()), None)
    if not target:
        raise HTTPException(status_code=404, detail=f"Page '{page_name}' not found")
    frames = [
        {"id": c["id"], "name": c["name"]}
        for c in target.get("children", [])
        if c.get("type") in ("FRAME", "COMPONENT")
    ]
    return {"frames": frames, "page_id": target["id"]}


@app.get("/api/figma/images")
async def get_figma_images(node_ids: str = Query(...)):
    token = os.environ.get("FIGMA_ACCESS_TOKEN")
    if not token:
        raise HTTPException(status_code=400, detail="FIGMA_ACCESS_TOKEN not set")
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.get(
            f"https://api.figma.com/v1/images/{FIGMA_FILE_KEY}",
            params={"ids": node_ids, "format": "png", "scale": "1"},
            headers={"X-Figma-Token": token},
        )
    data = resp.json()
    return {"images": data.get("images", {}), "err": data.get("err")}


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=True)
